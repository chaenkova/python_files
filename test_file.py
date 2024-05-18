from pypdf import PdfReader
from openpyxl import load_workbook
import csv

import os
import time
import requests
import zipfile


def download_file(url, filename):
    content = requests.get(url=url).content
    if not os.path.exists("tmp"):
        os.mkdir("tmp")

    with open(os.path.join("./tmp", filename), 'wb') as file:
        file.write(content)


download_file("https://github.com/qa-guru/qa_guru_python_9_7_files/raw/master/tmp/file_example_XLSX_50.xlsx","xlsx.xlsx")
time.sleep(4)

CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
if not os.path.exists("resources"):
    os.mkdir("resources")

RESOURCE_DIR = os.path.join(CURRENT_DIR, "resources")

with zipfile.ZipFile(os.path.join(RESOURCE_DIR, "zip.zip"), 'w') as zf:
    for file in ['xlsx.xlsx', 'pdf.pdf', 'csv.csv']:
        add_file = os.path.join("./tmp",file)
        zf.write(add_file, os.path.basename(add_file))

def test_pdf_file():
    with zipfile.ZipFile("resources/zip.zip", "r") as zip_file:
        page = PdfReader(zip_file.open('pdf.pdf'))
        assert 'Python Testing with pytest' in page.pages[1].extract_text()

def test_xlsx_file():
    with zipfile.ZipFile("resources/zip.zip", "r") as zip_file:
        page = load_workbook(zip_file.open('xlsx.xlsx')).active
        assert 'Mara' in page.cell(row=3, column=2).value

def test_csv_file():
    with zipfile.ZipFile("resources/zip.zip", "r") as zip_file:
        page = zip_file.open('csv.csv').read().decode('cp1251')
        csvreader = list(csv.reader(page.splitlines()))
        second_row = csvreader[1]
        assert 'дисплей' in second_row[0]